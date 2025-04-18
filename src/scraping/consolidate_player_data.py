import pandas as pd
import os
import sys
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

# Directorios
DATA_DIR = "data/player/player_details/"
VISUALS_DIR = "data/player/visuals/"  # ✅ Carpeta donde están las imágenes
OUTPUT_DIR = "data/player/consolidated/"

# Asegurar que el directorio de salida existe
os.makedirs(OUTPUT_DIR, exist_ok=True)

def consolidate_player_data(player_name):
    """
    Consolida todos los datos del jugador en un único archivo Excel en la carpeta de consolidación,
    incluyendo la imagen del jugador y el heatmap en hojas separadas.
    """
    player_filename = player_name.replace(" ", "_")
    output_file = os.path.join(OUTPUT_DIR, f"{player_filename}.xlsx")
    
    # Archivos CSV del jugador
    stat_files = {
        "Resumen": f"detalles_{player_filename}.csv",
        "Attacking": f"{player_filename}_Attacking.csv",
        "Passing": f"{player_filename}_Passing.csv",
        "Defending": f"{player_filename}_Defending.csv",
        "Other (per game)": f"{player_filename}_Other (per game).csv",
        "Cards": f"{player_filename}_Cards.csv",
    }
    
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for sheet_name, file_name in stat_files.items():
            file_path = os.path.join(DATA_DIR, file_name)
            if os.path.exists(file_path):
                df = pd.read_csv(file_path)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                print(f"✅ Agregada hoja: {sheet_name}")
            else:
                print(f"⚠ Archivo no encontrado: {file_path}")
    
    # Abrir el archivo para agregar imágenes en hojas separadas
    wb = load_workbook(output_file)

    # 📌 **Agregar imagen del jugador en una hoja nueva**
    player_image_file = os.path.join(VISUALS_DIR, f"{player_filename}.png")  # ✅ Carpeta corregida
    if os.path.exists(player_image_file):
        ws = wb.create_sheet("Imagen del Jugador")
        img = Image(player_image_file)
        ws.add_image(img, "A1")
        print("✅ Imagen del jugador agregada en hoja separada")
    else:
        print("⚠ No se encontró la imagen del jugador")

    # 📌 **Agregar imagen del heatmap en una hoja nueva**
    heatmap_file = os.path.join(VISUALS_DIR, f"heatmap_{player_filename}.png")  # ✅ Carpeta corregida
    if os.path.exists(heatmap_file):
        ws = wb.create_sheet("Mapa de Calor")
        img = Image(heatmap_file)
        ws.add_image(img, "A1")
        print("✅ Mapa de calor agregado en hoja separada")
    else:
        print("⚠ No se encontró el mapa de calor")

    # Guardar el archivo Excel con las imágenes en hojas separadas
    wb.save(output_file)
    
    print(f"📂 Archivo generado en: {output_file}")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("❌ Debes proporcionar el nombre del jugador.")
        sys.exit(1)
    
    player_name = sys.argv[1]
    consolidate_player_data(player_name)


# python -m src.scraping.consolidate_player_data  "Cristian Bernardi"